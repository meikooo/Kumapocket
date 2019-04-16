import random
import use
from openpyxl import load_workbook, workbook

member = load_workbook(filename=r"F:\\个人文档\\code\\Kumapocket\\resource\\20180927ribaomember.xlsx")
#读取名字为Sheet1的表
sheet = member.get_sheet_by_name("Sheet1")

#sheet = member.get_sheet_by_name(use wb["Sheet1"])

#sheet = workbook.worksheets[0]
#用于存储数据的数组
data= []
row_num = 1
while row_num <= 1 :
    data.append(sheet.cell(row=row_num, column=1).value)
    row_num = row_num + 1
#print(data)
slice = random.sample(data, 1)

print("恭喜以下日报鹅获得礼包！")
print(slice)